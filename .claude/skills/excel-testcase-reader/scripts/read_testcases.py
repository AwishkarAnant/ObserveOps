"""Read an Excel workbook of test cases and emit a Markdown report.

Expected columns (matched case-insensitively, whitespace-collapsed):
    - Title
    - Step Action
    - Step Expected Result

Usage:
    python read_testcases.py <input.xlsx> [-o <output.md>] [--sheet <name>]
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write(
        "openpyxl is not installed. Run: pip install openpyxl\n"
    )
    sys.exit(2)


REQUIRED_COLUMNS = ["Title", "Step Action", "Step Expected Result"]


def normalize(s: object) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def find_column_map(header_row: tuple) -> dict[str, int]:
    """Map required column names to their 0-based index in the header row."""
    normalized_headers = [normalize(c) for c in header_row]
    mapping: dict[str, int] = {}
    for col in REQUIRED_COLUMNS:
        key = normalize(col)
        if key in normalized_headers:
            mapping[col] = normalized_headers.index(key)
    return mapping


def cell_to_md(value: object) -> str:
    if value is None:
        return "_(empty)_"
    text = str(value).strip()
    if not text:
        return "_(empty)_"
    return text


def build_report(input_path: Path, sheet_name: str, rows: list[dict]) -> str:
    lines: list[str] = []
    lines.append(f"# Test Cases — {input_path.name}")
    lines.append("")
    lines.append(f"**Total test cases:** {len(rows)}")
    lines.append("")
    lines.append(f"**Source sheet:** {sheet_name}")
    lines.append("")
    lines.append("---")
    lines.append("")
    for i, row in enumerate(rows, start=1):
        lines.append(f"## {i}. {cell_to_md(row['Title'])}")
        lines.append("")
        lines.append(f"- **Step Action:** {cell_to_md(row['Step Action'])}")
        lines.append(
            f"- **Step Expected Result:** {cell_to_md(row['Step Expected Result'])}"
        )
        lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def read_testcases(
    input_path: Path, sheet_name: str | None = None
) -> tuple[str, list[dict], list[str]]:
    wb = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    all_sheets = wb.sheetnames
    if sheet_name is None:
        sheet_name = all_sheets[0]
    if sheet_name not in all_sheets:
        raise ValueError(
            f"Sheet '{sheet_name}' not found. Available: {all_sheets}"
        )
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = next(rows_iter)
    except StopIteration:
        return sheet_name, [], all_sheets

    col_map = find_column_map(header)
    missing = [c for c in REQUIRED_COLUMNS if c not in col_map]
    if missing:
        found = [str(h).strip() if h is not None else "" for h in header]
        raise ValueError(
            "Missing required columns: "
            + ", ".join(missing)
            + f". Found headers: {found}"
        )

    rows: list[dict] = []
    for raw in rows_iter:
        if raw is None:
            continue
        if all(
            v is None or (isinstance(v, str) and not v.strip()) for v in raw
        ):
            continue
        rows.append(
            {
                col: raw[idx] if idx < len(raw) else None
                for col, idx in col_map.items()
            }
        )
    return sheet_name, rows, all_sheets


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, help="Path to the .xlsx file")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=None,
        help="Output markdown path (default: <input-stem>_testcases.md next to input)",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="Sheet name to read (default: first sheet)",
    )
    args = parser.parse_args(argv)

    if not args.input.exists():
        sys.stderr.write(f"Input not found: {args.input}\n")
        return 1
    if args.input.suffix.lower() != ".xlsx":
        sys.stderr.write(
            f"Expected .xlsx, got {args.input.suffix}. "
            "openpyxl does not read legacy .xls files.\n"
        )
        return 1

    try:
        sheet_name, rows, all_sheets = read_testcases(args.input, args.sheet)
    except ValueError as e:
        sys.stderr.write(f"Error: {e}\n")
        return 1

    output_path = args.output or args.input.with_name(
        f"{args.input.stem}_testcases.md"
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    report = build_report(args.input, sheet_name, rows)
    output_path.write_text(report, encoding="utf-8")

    print(f"Wrote {len(rows)} test cases to {output_path}")
    if len(all_sheets) > 1 and args.sheet is None:
        other = [s for s in all_sheets if s != sheet_name]
        print(
            f"Note: workbook has other sheets ({other}). "
            f"Rerun with --sheet <name> to pick a different one."
        )
    return 0


if __name__ == "__main__":
    sys.exit(main())
